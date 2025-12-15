"""
Utilidad para importar archivos Excel y CSV.
Maneja la l√≥gica de procesamiento de archivos, validaci√≥n de datos y conversi√≥n.
Incluye mapeo inteligente de columnas para m√°xima flexibilidad.
"""

import openpyxl
import csv
import io
import re
from typing import List, Dict, Tuple, Optional


def normalizar_nombre_columna(nombre: str) -> str:
    """
    Normalizar nombre de columna para comparaci√≥n:
    - Quitar tildes y caracteres especiales
    - Convertir a min√∫sculas
    - Quitar espacios y caracteres no alfanum√©ricos
    """
    if not nombre:
        return ""
    
    # Convertir a string por si acaso
    nombre = str(nombre)
    
    # Primero normalizar con unicodedata para manejar mejor los caracteres especiales
    import unicodedata
    try:
        # Normalizar a NFD (Normalization Form Canonical Decomposition)
        # Esto separa los caracteres con tilde de sus acentos
        nombre_nfd = unicodedata.normalize('NFD', nombre)
        # Quitar los caracteres de combinaci√≥n (los acentos)
        nombre_sin_tildes = ''.join(c for c in nombre_nfd if unicodedata.category(c) != 'Mn')
    except:
        # Si falla, usar el m√©todo manual
        reemplazos = {
            '√°': 'a', '√©': 'e', '√≠': 'i', '√≥': 'o', '√∫': 'u',
            '√Å': 'a', '√â': 'e', '√ç': 'i', '√ì': 'o', '√ö': 'u',
            '√±': 'n', '√ë': 'n', '√º': 'u', '√ú': 'u'
        }
        nombre_sin_tildes = nombre
        for orig, reempl in reemplazos.items():
            nombre_sin_tildes = nombre_sin_tildes.replace(orig, reempl)
    
    # Convertir a min√∫sculas y quitar espacios/caracteres especiales
    nombre_normalizado = re.sub(r'[^a-z0-9]', '', nombre_sin_tildes.lower())
    
    nombre_normalizado = re.sub(r'[^a-z0-9]', '', nombre_sin_tildes.lower())
    
    return nombre_normalizado

def parse_date(date_val) -> Optional[object]:
    """
    Intenta convertir un valor a objeto date de Python.
    Soporta strings (YYYY-MM-DD, DD/MM/YYYY) y objetos datetime/date.
    """
    if not date_val:
        return None
        
    from datetime import datetime, date
    
    # Si ya es date o datetime
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
        
    date_str = str(date_val).strip()
    if not date_str:
        return None
        
    # Intentar varios formatos
    formats = [
        '%Y-%m-%d',       # 2023-12-25
        '%d/%m/%Y',       # 25/12/2023
        '%d-%m-%Y',       # 25-12-2023
        '%Y/%m/%d',       # 2023/12/25
        '%d/%m/%y',       # 25/12/23
        '%Y-%m-%d %H:%M:%S' # Timestamp string
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
            
    return None


def mapear_columnas(headers: List[str]) -> Dict[str, str]:
    """
    Mapear headers del archivo a nombres est√°ndar de columnas.
    Retorna un diccionario: {nombre_estandar: nombre_real_en_archivo}
    """
    # Definir todas las variaciones posibles para cada campo
    mapeo_campos = {
        'nombre': [
            'nombre', 'name', 'nombrecompleto', 'nombreyfamiliar', 
            'nombrepersona', 'titulo', 'persona', 'fullname',
            'nombres', 'nombreyapellido', 'nombreyapellidos'
        ],
        'apellido': [
            'apellido', 'lastname', 'surname', 'apellidopersona',
            'apellidodelconyuge', 'apellidodel', 'apellidos'
        ],
        'direccion': [
            'direccion', 'address', 'ubicacion', 'domicilio', 
            'calle', 'residencia', 'descripcion', 'direccionderesidencia',
            'direccionresidencia', 'direccioncompl', 'direccionderesidenciacompl',
            'direcci', 'direc', 'domicili', 'residenci',
            'direcciondecasa', 'direccionhabitacion', 'lugar', 'dondevive'
        ],
        'codigo_postal': [
            'codigopostal', 'cp', 'zip', 'zipcode', 'postalcode', 
            'codigop', 'postal', 'zona'
        ],
        'telefono': [
            'telefono', 'phone', 'tel', 'celular', 'movil', 
            'telefonopersonal', 'telefonocelular', 'contacto',
            'numerodetelefono', 'numerotel', 'telno', 'numtelefono',
            'numerodtel', 'telef', 'numerodetlf', 'numerodetel',
            'numerodeteleno', 'numdetel', 'teleno',
            'telefonomovil', 'telefonocasa', 'whatsapp'
        ],
        'edad': [
            'edad', 'age', 'anos', 'years', 'edadactual'
        ],
        'estado_civil': [
            'estadocivil', 'maritalstatus', 'estadomarital', 
            'civilstatus', 'estatus', 'estadocivilactual',
            'estado', 'situacioncivil', 'condicioncivil', 'civil'
        ],
        'num_hijos': [
            'numerohijos', 'numerodehijos', 'cantidadhijos', 'nohijos',
            'children', 'cuantoshijos', 'cuantos', 'numhijos', 'nohijos'
        ],
        'edades_hijos': [
            'edadeshijos', 'edadesdehijos', 'edadeshijo', 
            'childrenages', 'edadesdeloshijos'
        ],
        'nombre_conyuge': [
            'nombreconyuge', 'conyuge', 'esposo', 'esposa', 
            'spouse', 'nombreesposo', 'nombreesposa', 'pareja',
            'spousename', 'nombredeesposo', 'nombredeesposa',
            'nombredelconyuge', 'nombredeconyuge', 'nombredelesposo',
            'nombrepareja'
        ],
        'apellido_conyuge': [
            'apellidoconyuge', 'apellidodelconyuge', 'apellidodeesposo',
            'apellidodeesposa', 'apellidodel', 'surnamesp spouse'
        ],
        'edad_conyuge': [
            'edadconyuge', 'edadesposo', 'edadesposa', 
            'spouseage', 'edaddeesposo', 'edaddeesposa'
        ],
        'telefono_conyuge': [
            'telefonoconyuge', 'telefonoesposo', 'telefonoesposa',
            'spousephone', 'telconyuge', 'celularconyuge',
            'telefonodeesposo', 'telefonodeesposa',
            'numerodetelfonodelconyuge', 'numerotelefonoconyuge',
            'numerodelconyuge', 'numerodetelefonodelconyuge',
            'telefonodelconyuge', 'numtelconyuge'
        ],
        'trabajo_conyuge': [
            'trabajoconyuge', 'ocupacionconyuge', 'empleoconyuge',
            'spousework', 'spousejob', 'trabajoesposo', 'trabajoesposa',
            'profesionesposo', 'profesionesposa', 'ocupacionesposo'
        ],
        'fecha_matrimonio': [
            'fechamatrimonio', 'matrimonio', 'fechadeboda', 
            'marriagedate', 'fechacasamiento', 'aniomatrimonio'
        ],
        'ocupacion': [
            'ocupacion', 'trabajo', 'empleo', 'profession', 
            'job', 'profesion', 'oficio', 'carrera',
            'profesionoficio', 'prof', 'profesionooficio'
        ],

        'email_conyuge': [
            'emailconyuge', 'correoconyuge', 'correodelconyuge', 
            'emaildelconyuge', 'correoelectronicodelconyuge', 
            'correoelectronicoconyuge', 'mailconyuge'
        ],
        'email': [
            'email', 'correo', 'correoelectronico', 'mail', 
            'emailaddress', 'electronico', 'correoe', 'emailpersonal'
        ],
        'responsable': [
            'responsable', 'registradopor', 'capturo', 
            'registrador', 'ingresadopor', 'responsible'
        ],
        'notas': [
            'notas', 'observaciones', 'comentarios', 'notes', 
            'observacion', 'comentario', 'remarks', 'adicional'
        ]
    }
    
    # Normalizar headers
    headers_normalizados = {normalizar_nombre_columna(h): h for h in headers if h}
    
    # Encontrar coincidencias
    resultado = {}
    import difflib
    
    # 1. Intentar coincidencia exacta primero
    for campo_estandar, variaciones in mapeo_campos.items():
        match_found = False
        for variacion in variaciones:
            variacion_normalizada = normalizar_nombre_columna(variacion)
            if variacion_normalizada in headers_normalizados:
                resultado[campo_estandar] = headers_normalizados[variacion_normalizada]
                match_found = True
                break
        
        # 2. Si no hay coincidencia exacta, usar fuzzy matching
        if not match_found:
            # Crear lista de posibles nombres de columna normalizados del archivo
            columnas_archivo = list(headers_normalizados.keys())
            
            # Buscar la mejor coincidencia para cada variaci√≥n
            mejores_coincidencias = []
            for variacion in variaciones:
                variacion_normalizada = normalizar_nombre_columna(variacion)
                matches = difflib.get_close_matches(variacion_normalizada, columnas_archivo, n=1, cutoff=0.8)
                if matches:
                    mejores_coincidencias.append((matches[0], variacion))
            
            # Si encontramos alguna coincidencia fuzzy
            if mejores_coincidencias:
                # Usar la primera (podr√≠amos mejorar esto eligiendo la de mayor score)
                mejor_match = mejores_coincidencias[0][0]
                resultado[campo_estandar] = headers_normalizados[mejor_match]
                print(f"‚ú® Coincidencia inteligente: '{headers_normalizados[mejor_match]}' -> {campo_estandar}")
    
    return resultado


def obtener_valor_flexible(fila: Dict, mapeo: Dict[str, str], campo: str) -> str:
    """
    Obtener valor de una fila usando el mapeo flexible.
    """
    nombre_columna_real = mapeo.get(campo)
    if nombre_columna_real:
        return str(fila.get(nombre_columna_real, '')).strip()
    return ''


def process_excel_file(archivo) -> Tuple[List[Dict], List[str]]:
    """
    Procesar archivo Excel (.xlsx, .xls) con detecci√≥n autom√°tica de fila de headers.
    
    Args:
        archivo: FileStorage object del archivo
        
    Returns:
        Tuple con (lista de filas procesadas, lista de errores)
    """
    filas_datos = []
    errores = []
    
    try:
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
        
        # Detectar qu√© fila tiene los headers
        # Los headers suelen tener texto, no n√∫meros, y en m√∫ltiples columnas
        header_row = 1
        max_rows_to_check = min(5, sheet.max_row)
        
        for row_num in range(1, max_rows_to_check + 1):
            row_values = [cell.value for cell in sheet[row_num]]
            
            # Contar cu√°ntas celdas tienen texto (no None, no n√∫meros puros)
            text_cells = 0
            for val in row_values:
                if val and isinstance(val, str) and len(val.strip()) > 0:
                    text_cells += 1
            
            # Si esta fila tiene 3+ celdas con texto, probablemente son headers
            if text_cells >= 3:
                # Verificar que no sea una fila de datos mirando si tiene palabras comunes de headers
                row_text = ' '.join([str(v).lower() for v in row_values if v])
                header_keywords = ['nombre', 'name', 'direccion', 'address', 'telefono', 'phone', 
                                 'edad', 'age', 'email', 'correo']
                
                if any(keyword in row_text for keyword in header_keywords):
                    header_row = row_num
                    break
        
        # Extraer headers de la fila detectada
        headers = [cell.value for cell in sheet[header_row]]
        
        print(f"üîç Headers detectados en fila {header_row}")
        
        # Procesar cada fila de datos (despu√©s de los headers)
        for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            fila_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    fila_dict[headers[i]] = str(value) if value else ''
            
            # Solo agregar si tiene alg√∫n dato
            if any(fila_dict.values()):
                fila_dict['_row_num'] = row_num
                filas_datos.append(fila_dict)
        
    except Exception as e:
        errores.append(f'Error procesando Excel: {str(e)}')
    
    return filas_datos, errores


def process_csv_file(archivo) -> Tuple[List[Dict], List[str]]:
    """
    Procesar archivo CSV
    
    Args:
        archivo: FileStorage object del archivo
        
    Returns:
        Tuple con (lista de filas procesadas, lista de errores)
    """
    filas_datos = []
    errores = []
    
    try:
        # Leer el archivo como bytes
        content_bytes = archivo.read()
        archivo_contenido = None
        
        # Intentar decodificar con diferentes encodings
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for encoding in encodings:
            try:
                archivo_contenido = content_bytes.decode(encoding)
                print(f"‚úÖ Archivo CSV decodificado con {encoding}")
                break
            except UnicodeDecodeError:
                continue
                
        if archivo_contenido is None:
            errores.append('Error de codificaci√≥n: El archivo no es UTF-8 ni Latin-1 v√°lido.')
            return filas_datos, errores
        
        # Validar m√≠nimo de l√≠neas
        lineas = archivo_contenido.strip().split('\n')
        if len(lineas) < 2:
            errores.append('El archivo debe tener al menos una fila de encabezados y una fila de datos')
            return filas_datos, errores
        
        # Detectar delimitador
        try:
            # Tomar una muestra representativa (primeras 5 l√≠neas)
            sample = '\n'.join(lineas[:5])
            dialect = csv.Sniffer().sniff(sample)
            delimiter = dialect.delimiter
        except:
            # Fallback si falla la detecci√≥n
            # Contar comas vs punto y coma en la primera l√≠nea
            header = lineas[0]
            delimiter = ';' if header.count(';') > header.count(',') else ','
            
        print(f"üîç Delimitador CSV detectado: '{delimiter}'")
        
        print(f"üîç Delimitador CSV detectado: '{delimiter}'")
        
        # Detectar fila de headers
        header_row_index = 0
        header_keywords = ['nombre', 'name', 'direccion', 'address', 'telefono', 'phone', 
                         'edad', 'age', 'email', 'correo', 'apellido']
        
        # Buscar la primera fila que parezca tener headers
        csv_reader_temp = csv.reader(io.StringIO(archivo_contenido), delimiter=delimiter)
        rows_temp = list(csv_reader_temp)
        
        for i, row in enumerate(rows_temp):
            # Convertir fila a string para buscar keywords
            row_text = ' '.join([str(v).lower() for v in row if v])
            
            # Si encontramos al menos 2 keywords, asumimos que es el header
            matches = sum(1 for k in header_keywords if k in row_text)
            if matches >= 2:
                header_row_index = i
                print(f"üîç Headers detectados en fila CSV {i+1}")
                break
        
        # Reconstruir contenido desde la fila de headers
        # Usamos rows_temp[header_row_index:]
        if not rows_temp:
             return filas_datos, errores

        # Obtener los headers de la fila detectada
        headers = rows_temp[header_row_index]
        
        # Procesar datos
        for row_num, row_values in enumerate(rows_temp[header_row_index+1:], start=header_row_index+2):
            if not row_values: continue
            
            # Crear diccionario manual para evitar problemas con DictReader y l√≠neas vac√≠as previas
            fila = {}
            for i, val in enumerate(row_values):
                if i < len(headers):
                    header = headers[i]
                    if header: # Solo si el header tiene nombre
                        fila[header] = val
            
            if any(fila.values()): # Solo si tiene datos
                fila['_row_num'] = row_num
                filas_datos.append(fila)
    
    except Exception as e:
        errores.append(f'Error procesando CSV: {str(e)}')
    
    return filas_datos, errores


def extract_person_data(fila: Dict, row_num: int = None, mapeo: Dict[str, str] = None) -> Tuple[Dict, List[str]]:
    """
    Extraer y validar datos de una persona desde una fila usando mapeo inteligente.
    
    Args:
        fila: Diccionario con los datos de la fila
        row_num: N√∫mero de fila (opcional, para mensajes de error)
        mapeo: Diccionario de mapeo de columnas (opcional)
        
    Returns:
        Tuple con (diccionario de datos de persona, lista de errores)
    """
    errores = []
    row_num = row_num or fila.get('_row_num', 0)
    
    # Si no hay mapeo, crearlo
    if mapeo is None:
        headers = list(fila.keys())
        mapeo = mapear_columnas(headers)
    
    # Extraer nombre completo
    nombre_completo = obtener_valor_flexible(fila, mapeo, 'nombre')
    
    if not nombre_completo:
        errores.append(f'Fila {row_num}: Nombre es obligatorio')
        return {}, errores
    
    # Intentar obtener apellido de una columna separada
    apellido_separado = obtener_valor_flexible(fila, mapeo, 'apellido')
    
    if apellido_separado:
        # Si hay columna de apellido separada, usarla
        nombre = nombre_completo
        apellido = apellido_separado
    else:
        # Si no, separar del nombre completo
        partes_nombre = nombre_completo.split(' ', 1)
        nombre = partes_nombre[0]
        apellido = partes_nombre[1] if len(partes_nombre) > 1 else ''
    
    # Extraer otros campos usando el mapeo
    direccion = obtener_valor_flexible(fila, mapeo, 'direccion')
    codigo_postal = obtener_valor_flexible(fila, mapeo, 'codigo_postal')
    telefono = obtener_valor_flexible(fila, mapeo, 'telefono')
    edad_str = obtener_valor_flexible(fila, mapeo, 'edad')
    estado_civil = obtener_valor_flexible(fila, mapeo, 'estado_civil')
    hijos_str = obtener_valor_flexible(fila, mapeo, 'num_hijos') or '0'
    edades_hijos = obtener_valor_flexible(fila, mapeo, 'edades_hijos')
    
    # Nombre del c√≥nyuge - combinar con apellido si existe
    nombre_conyuge = obtener_valor_flexible(fila, mapeo, 'nombre_conyuge')
    apellido_conyuge = obtener_valor_flexible(fila, mapeo, 'apellido_conyuge')
    if apellido_conyuge and nombre_conyuge:
        nombre_conyuge = f"{nombre_conyuge} {apellido_conyuge}"
    
    # Inferencia inteligente de estado civil
    if not estado_civil:
        if nombre_conyuge:
            estado_civil = 'Casado'
        else:
            estado_civil = 'Soltero'
    
    edad_conyuge_str = obtener_valor_flexible(fila, mapeo, 'edad_conyuge')
    telefono_conyuge = obtener_valor_flexible(fila, mapeo, 'telefono_conyuge')
    email_conyuge = obtener_valor_flexible(fila, mapeo, 'email_conyuge')
    telefono_conyuge = obtener_valor_flexible(fila, mapeo, 'telefono_conyuge')
    email_conyuge = obtener_valor_flexible(fila, mapeo, 'email_conyuge')
    trabajo_conyuge = obtener_valor_flexible(fila, mapeo, 'trabajo_conyuge')
    
    # Parsear fechas
    fecha_matrimonio_raw = obtener_valor_flexible(fila, mapeo, 'fecha_matrimonio')
    fecha_matrimonio = parse_date(fecha_matrimonio_raw)
    
    ocupacion = obtener_valor_flexible(fila, mapeo, 'ocupacion')
    email = obtener_valor_flexible(fila, mapeo, 'email')
    notas = obtener_valor_flexible(fila, mapeo, 'notas')
    
    # Convertir edad
    edad = None
    if edad_str:
        try:
            edad = int(float(edad_str))
        except ValueError:
            errores.append(f'Fila {row_num}: Edad "{edad_str}" no es un n√∫mero v√°lido')
    
    # Convertir edad del c√≥nyuge
    edad_conyuge = None
    if edad_conyuge_str:
        try:
            edad_conyuge = int(float(edad_conyuge_str))
        except ValueError:
            pass  # No es cr√≠tico
    
    # Convertir n√∫mero de hijos
    numero_hijos = 0
    if hijos_str:
        try:
            numero_hijos = int(float(hijos_str))
        except ValueError:
            errores.append(f'Fila {row_num}: N√∫mero de hijos "{hijos_str}" no es v√°lido')
    
    # Construir diccionario de datos
    persona_data = {
        'nombre': nombre,
        'apellido': apellido,
        'direccion': direccion,
        'codigo_postal': codigo_postal,
        'telefono': telefono,
        'edad': edad,
        'estado_civil': estado_civil,
        'numero_hijos': numero_hijos,
        'edades_hijos': edades_hijos,
        'nombre_conyuge': nombre_conyuge,
        'telefono_conyuge': telefono_conyuge,
        'email_conyuge': email_conyuge,
        'edad_conyuge': edad_conyuge,
        'trabajo_conyuge': trabajo_conyuge,
        'fecha_matrimonio': fecha_matrimonio,
        'ocupacion': ocupacion,
        'email': email,
        'notas': notas
    }
    
    return persona_data, errores


def process_import_file(archivo, filename: str) -> Tuple[List[Dict], List[str], str, List[str]]:
    """
    Procesar archivo de importaci√≥n (Excel o CSV) con mapeo inteligente de columnas.
    
    Args:
        archivo: FileStorage object
        filename: Nombre del archivo
        
    Returns:
        Tuple con (lista de datos de personas, lista de errores, tipo de archivo, columnas faltantes)
    """
    filename_lower = filename.lower()
    personas_data = []
    all_errors = []
    file_type = 'unknown'
    columnas_faltantes = []
    
    # Procesar seg√∫n el tipo de archivo
    if filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        file_type = 'excel'
        filas, errores = process_excel_file(archivo)
        all_errors.extend(errores)
    elif filename_lower.endswith('.csv'):
        file_type = 'csv'
        filas, errores = process_csv_file(archivo)
        all_errors.extend(errores)
    else:
        all_errors.append('Formato de archivo no soportado. Use .xlsx, .xls o .csv')
        return personas_data, all_errors, file_type, columnas_faltantes
    
    if not filas:
        all_errors.append('No se encontraron datos en el archivo')
        return personas_data, all_errors, file_type, columnas_faltantes
    
    # Crear mapeo de columnas una vez usando la primera fila
    headers = list(filas[0].keys()) if filas else []
    mapeo = mapear_columnas(headers)
    
    # Log de columnas encontradas (√∫til para debugging)
    print(f"üìä Mapeo de columnas detectado:")
    for campo_std, columna_real in mapeo.items():
        print(f"  ‚úì {campo_std:20s} <- '{columna_real}'")
    
    campos_importantes = ['nombre', 'telefono', 'direccion', 'email', 'estado_civil']
    for campo in campos_importantes:
        if campo not in mapeo:
            columnas_faltantes.append(campo)
    
    if columnas_faltantes:
        print(f"  ‚ö†Ô∏è  Columnas no encontradas: {', '.join(columnas_faltantes)}")
    
    # Extraer datos de personas de cada fila usando el mismo mapeo
    for fila in filas:
        persona_data, errores = extract_person_data(fila, mapeo=mapeo)
        if persona_data:
            personas_data.append(persona_data)
        all_errors.extend(errores)
    
    return personas_data, all_errors, file_type, columnas_faltantes
